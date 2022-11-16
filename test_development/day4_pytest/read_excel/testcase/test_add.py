import pytest
from test_development.day4_pytest.read_excel.func.operation import test_add
import openpyxl
import os
# 判断表格是否存在
def get_excel():
    # if os.path.exists("../data/params.xlsx"):
    #     # 如果存在就获取
    #     wb = openpyxl.load_workbook("../data/params.xlsx")
    # else:
    #     # 不存在就新建空白表格
    #     wb = openpyxl.Workbook()

    # 获取工作薄
    wb = openpyxl.load_workbook("/Users/zxmac/Desktop/workspace/project/hogwarts/test_development/day4_pytest/"
                                "read_excel/data/params.xlsx")
    # 读取工作表
    sheet = wb.active
    # 输出A1的值
    print(sheet['A1'].value, "aaaaa")
    # 输出A1到C3的所有值
    datalist = sheet["A1":"C3"]
    # 打印出类型和所有值列表
    print(type(datalist), datalist, "bbbbb")

    # 定义一个列表用来放从元组中遍历出来的列表位置对应的值，举例：sheet['A1'].value
    list1 = []
    for data in datalist:
        list2 = []
        print(data, "ccccc")
        for dat in data:
            print(dat, "ddddd")
            # 遍历元组，出来的列表再次遍历，获取到的每个数据所在表格的键，然后把这个键对应的值放入list2这个列表中
            list2.append(dat.value)
        print(type(list2), list2, "eeeee")
        # 再把每个列表组合到一个大列表中
        list1.append(list2)
    print(type(list1), list1, "fffff")
    return list1

    # wb.save("/Users/zxmac/Desktop/workspace/project/hogwarts/test_development/day4_pytest/read_excel/data/params.xlsx")
    # wb.close()


# 测试用例

class TestWithExcel:
    # 执行测试用例时，命名三个参数，和导入已经return给get_excel()的列表数据，参数和数据只要位置能够对应，会自动匹配
    @pytest.mark.parametrize('x, y, expect', get_excel())
    def test_my_add(self, x, y, expect):
        assert test_add(int(x), int(y)) == int(expect)

