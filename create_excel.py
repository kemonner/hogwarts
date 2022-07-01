from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 首先实例化Workbook
wb = Workbook()
# 定义变量存放excel文件名
dest_filename = 'empty_book.xlsx'
# 通过active调用正在运行的工作表，赋值给ws1
ws1 = wb.active
ws1.title = "range names"
ws1["A1"] = "ID"
ws1["B1"] = "姓名"
ws1["C1"] = "职位"
ws1["D1"] = "部门"
ws1["E1"] = "地区"
for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

w4 = wb.create_sheet(title="test_sheet")
for i in range(1, 30):
    w4.cell(column=1, row=i, value="test")

wb.save(filename=dest_filename)
