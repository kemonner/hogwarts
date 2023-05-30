import openpyxl
import os

# 获取工作薄
# 检测到xlsx文件存在就获取
if os.path.exists('/TestDevelopment/day4_pytest/testxlsx1.xlsx'):
    new_wb = openpyxl.load_workbook('/TestDevelopment/day4_pytest/'
                                    'testxlsx1.xlsx')
# 不存在的话就创建一个空白excl
else:
    new_wb = openpyxl.Workbook()


# 读取工作表
sheet = new_wb.active

# 读取单个单元格
a1 = sheet['A1'].value
print(a1)
a3 = sheet.cell(column=1, row=3).value
print(a3)

# 获取多个单元格
cells = sheet['A1':'C3']
print(type(cells), cells)

# 保存并关闭
new_wb.save('/Users/zxmac/Desktop/workspace/project/hogwarts/TestDevelopment/day4_pytest/testxlsx1.xlsx')
new_wb.close()
